"""
Reusable Excel/CSV export service.

Usage:
    from app.services.export_service import ExportService
    return ExportService.to_excel_response(rows, columns, filename="products")
    return ExportService.to_csv_response(rows, columns, filename="products")
"""
import io
import csv
from datetime import datetime
from typing import Any

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExportService:

    @staticmethod
    def to_excel_response(
        rows: list[dict],
        columns: dict[str, str],
        filename: str = "export",
        sheet_name: str = "Datos",
    ) -> StreamingResponse:
        """
        Generate a styled Excel file and return as StreamingResponse.

        Args:
            rows: list of dicts with raw data
            columns: ordered dict mapping field_key -> display_header
            filename: base name for the download file
            sheet_name: name of the Excel sheet
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # --- Header style ---
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = list(columns.values())
        keys = list(columns.keys())

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, key in enumerate(keys, 1):
                value = row_data.get(key, "")
                # Handle nested keys like "client.name"
                if "." in key:
                    parts = key.split(".")
                    value = row_data
                    for part in parts:
                        if isinstance(value, dict):
                            value = value.get(part, "")
                        else:
                            value = getattr(value, part, "")
                
                # Format datetime objects
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M")
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-width columns
        for col_idx, key in enumerate(keys, 1):
            max_length = len(headers[col_idx - 1])
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 3, 50)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}_{timestamp}.xlsx"'
            },
        )

    @staticmethod
    def to_csv_response(
        rows: list[dict],
        columns: dict[str, str],
        filename: str = "export",
    ) -> StreamingResponse:
        """Generate a CSV file and return as StreamingResponse."""
        buffer = io.StringIO()
        keys = list(columns.keys())
        headers = list(columns.values())

        writer = csv.writer(buffer)
        writer.writerow(headers)

        for row_data in rows:
            row_values = []
            for key in keys:
                value = row_data.get(key, "")
                if "." in key:
                    parts = key.split(".")
                    value = row_data
                    for part in parts:
                        if isinstance(value, dict):
                            value = value.get(part, "")
                        else:
                            value = getattr(value, part, "")
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M")
                row_values.append(value)
            writer.writerow(row_values)

        buffer.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}_{timestamp}.csv"'
            },
        )

    @staticmethod
    def model_to_dict(obj: Any, keys: list[str]) -> dict:
        """Convert an SQLAlchemy model instance to a dict, supporting nested keys."""
        result = {}
        for key in keys:
            if "." in key:
                parts = key.split(".")
                value = obj
                for part in parts:
                    value = getattr(value, part, None) if value else None
                result[key] = value if value is not None else ""
            else:
                result[key] = getattr(obj, key, "")
        return result
