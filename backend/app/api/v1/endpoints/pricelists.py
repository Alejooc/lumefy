from typing import Any, List
from io import BytesIO
from decimal import Decimal, InvalidOperation
import uuid
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from sqlalchemy import delete

from app.api import deps
from app.api.deps import get_db, PermissionChecker
from app.models.user import User
from app.models.pricelist import PriceList
from app.models.pricelist_item import PriceListItem
from app.models.pricelist_source_rule import PriceListSourceRule
from app.models.product import Product
from app.models.product_variant import ProductVariant
from app.models.integration import IntegrationSource
from app.schemas import pricelist as schemas
from app.core.audit import log_activity
from app.services.pricing import load_price_list_context, resolve_price, selected_external_values

router = APIRouter()


async def _get_pricelist(db: AsyncSession, list_id: str, current_user: User) -> PriceList:
    result = await db.execute(
        select(PriceList)
        .where(PriceList.id == list_id, PriceList.company_id == current_user.company_id)
        .options(selectinload(PriceList.items), selectinload(PriceList.source_rules))
    )
    pricelist = result.scalars().first()
    if not pricelist:
        raise HTTPException(status_code=404, detail="Lista de precios no encontrada")
    return pricelist


async def _validate_item_product(
    db: AsyncSession,
    item: schemas.PriceListItemCreate,
    company_id,
) -> tuple[Product, ProductVariant | None]:
    product = await db.scalar(
        select(Product).where(Product.id == item.product_id, Product.company_id == company_id)
    )
    if not product:
        raise HTTPException(status_code=400, detail="El producto no pertenece a la empresa")
    variant = None
    if item.variant_id:
        variant = await db.scalar(
            select(ProductVariant).where(
                ProductVariant.id == item.variant_id,
                ProductVariant.product_id == product.id,
                ProductVariant.company_id == company_id,
            )
        )
        if not variant:
            raise HTTPException(status_code=400, detail="La variante no pertenece al producto")
    return product, variant


async def _validate_source(db: AsyncSession, source_id, company_id) -> None:
    if source_id is None:
        return
    source = await db.scalar(
        select(IntegrationSource).where(
            IntegrationSource.id == source_id,
            IntegrationSource.company_id == company_id,
        )
    )
    if not source:
        raise HTTPException(status_code=400, detail="El origen de datos no pertenece a la empresa")


def _parse_excel_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    normalized = str(value).strip().replace("$", "").replace(" ", "")
    if not normalized:
        return None
    # Accept both 1234.56 and Colombian-style 1.234,56 values.
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return float(normalized)
    except (TypeError, ValueError, InvalidOperation):
        return None

@router.get("/", response_model=List[schemas.PriceList])
async def read_pricelists(
    skip: int = 0,
    limit: int = 100,
    type: str = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> Any:
    """
    Retrieve price lists.
    """
    query = select(PriceList).options(
        selectinload(PriceList.items), selectinload(PriceList.source_rules)
    ).where(
        PriceList.company_id == current_user.company_id
    )
    if type:
        query = query.where(PriceList.type == type)
    
    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    return result.scalars().all()

@router.post("/", response_model=schemas.PriceList)
async def create_pricelist(
    *,
    db: AsyncSession = Depends(get_db),
    pricelist_in: schemas.PriceListCreate,
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> Any:
    """
    Create new price list.
    """
    pricelist = PriceList(
        name=pricelist_in.name,
        type=pricelist_in.type,
        currency=pricelist_in.currency,
        active=pricelist_in.active,
        source_id=pricelist_in.source_id,
        pricing_mode=pricelist_in.pricing_mode,
        base_source=pricelist_in.base_source,
        adjustment_value=pricelist_in.adjustment_value,
        rounding_step=pricelist_in.rounding_step,
        min_margin_percent=pricelist_in.min_margin_percent,
        company_id=current_user.company_id
    )
    await _validate_source(db, pricelist_in.source_id, current_user.company_id)
    if pricelist_in.base_source.startswith("EXTERNAL") and not pricelist_in.source_id:
        raise HTTPException(status_code=400, detail="Selecciona el origen de datos para usar un precio externo")
    db.add(pricelist)
    await db.flush()

    if pricelist_in.source_rules and pricelist_in.type.value != "SALE":
        raise HTTPException(status_code=400, detail="Las reglas por proveedor solo aplican a listas de venta")
    for rule_in in pricelist_in.source_rules:
        await _validate_source(db, rule_in.source_id, current_user.company_id)

    for item_in in pricelist_in.items:
        await _validate_item_product(db, item_in, current_user.company_id)
        item = PriceListItem(
            pricelist_id=pricelist.id,
            product_id=item_in.product_id,
            variant_id=item_in.variant_id,
            min_quantity=item_in.min_quantity,
            price=item_in.price
        )
        db.add(item)
    
    for rule_in in pricelist_in.source_rules:
        db.add(PriceListSourceRule(
            pricelist_id=pricelist.id,
            company_id=current_user.company_id,
            **rule_in.model_dump(),
        ))

    await db.commit()
    refreshed = await db.execute(
        select(PriceList).where(
            PriceList.id == pricelist.id,
            PriceList.company_id == current_user.company_id,
        ).options(selectinload(PriceList.items), selectinload(PriceList.source_rules))
    )
    pricelist = refreshed.scalars().first()
    await log_activity(db, "CREATE", "PriceList", pricelist.id, current_user.id, current_user.company_id)
    return pricelist

@router.get("/{id}", response_model=schemas.PriceList)
async def read_pricelist(
    id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> Any:
    """
    Get price list by ID.
    """
    query = select(PriceList).where(
        PriceList.id == id,
        PriceList.company_id == current_user.company_id
    ).options(selectinload(PriceList.items), selectinload(PriceList.source_rules))
    result = await db.execute(query)
    pricelist = result.scalars().first()
    if not pricelist:
        raise HTTPException(status_code=404, detail="Price list not found")
    return pricelist


@router.post("/{id}/source-rules", response_model=schemas.PriceListSourceRule)
async def upsert_pricelist_source_rule(
    *,
    db: AsyncSession = Depends(get_db),
    id: str,
    rule_in: schemas.PriceListSourceRuleCreate,
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> PriceListSourceRule:
    pricelist = await _get_pricelist(db, id, current_user)
    if str(getattr(pricelist.type, "value", pricelist.type)) != "SALE":
        raise HTTPException(status_code=400, detail="Las reglas por proveedor solo aplican a listas de venta")
    await _validate_source(db, rule_in.source_id, current_user.company_id)
    rule = await db.scalar(
        select(PriceListSourceRule).where(
            PriceListSourceRule.pricelist_id == pricelist.id,
            PriceListSourceRule.source_id == rule_in.source_id,
        )
    )
    if rule is None:
        rule = PriceListSourceRule(
            pricelist_id=pricelist.id,
            company_id=current_user.company_id,
            **rule_in.model_dump(),
        )
        db.add(rule)
    else:
        for field, value in rule_in.model_dump().items():
            setattr(rule, field, value)
        rule.is_active = True
    await db.commit()
    await db.refresh(rule)
    return rule


@router.put("/{id}/source-rules/{rule_id}", response_model=schemas.PriceListSourceRule)
async def update_pricelist_source_rule(
    *,
    db: AsyncSession = Depends(get_db),
    id: str,
    rule_id: str,
    rule_in: schemas.PriceListSourceRuleUpdate,
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> PriceListSourceRule:
    pricelist = await _get_pricelist(db, id, current_user)
    if str(getattr(pricelist.type, "value", pricelist.type)) != "SALE":
        raise HTTPException(status_code=400, detail="Las reglas por proveedor solo aplican a listas de venta")
    rule = await db.scalar(
        select(PriceListSourceRule).where(
            PriceListSourceRule.id == rule_id,
            PriceListSourceRule.pricelist_id == pricelist.id,
        )
    )
    if not rule:
        raise HTTPException(status_code=404, detail="Regla de proveedor no encontrada")
    update_data = rule_in.model_dump(exclude_unset=True)
    if "source_id" in update_data:
        if not update_data["source_id"]:
            raise HTTPException(status_code=400, detail="Selecciona un proveedor")
        await _validate_source(db, update_data["source_id"], current_user.company_id)
        duplicate = await db.scalar(
            select(PriceListSourceRule.id).where(
                PriceListSourceRule.pricelist_id == pricelist.id,
                PriceListSourceRule.source_id == update_data["source_id"],
                PriceListSourceRule.id != rule.id,
            )
        )
        if duplicate:
            raise HTTPException(status_code=409, detail="Este proveedor ya tiene una regla en la lista")
    for field, value in update_data.items():
        setattr(rule, field, value)
    await db.commit()
    await db.refresh(rule)
    return rule


@router.delete("/{id}/source-rules/{rule_id}", status_code=204)
async def delete_pricelist_source_rule(
    *,
    db: AsyncSession = Depends(get_db),
    id: str,
    rule_id: str,
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> None:
    pricelist = await _get_pricelist(db, id, current_user)
    rule = await db.scalar(
        select(PriceListSourceRule).where(
            PriceListSourceRule.id == rule_id,
            PriceListSourceRule.pricelist_id == pricelist.id,
        )
    )
    if not rule:
        raise HTTPException(status_code=404, detail="Regla de proveedor no encontrada")
    await db.delete(rule)
    await db.commit()

@router.put("/{id}", response_model=schemas.PriceList)
async def update_pricelist(
    *,
    db: AsyncSession = Depends(get_db),
    id: str,
    pricelist_in: schemas.PriceListUpdate,
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> Any:
    """
    Update a price list.
    """
    query = select(PriceList).where(
        PriceList.id == id,
        PriceList.company_id == current_user.company_id
    )
    result = await db.execute(query)
    pricelist = result.scalars().first()
    if not pricelist:
        raise HTTPException(status_code=404, detail="Price list not found")
    
    update_data = pricelist_in.dict(exclude_unset=True)
    effective_type = update_data.get("type", pricelist.type)
    if str(getattr(effective_type, "value", effective_type)) != "SALE":
        existing_rule = await db.scalar(
            select(PriceListSourceRule.id).where(PriceListSourceRule.pricelist_id == pricelist.id)
        )
        if existing_rule:
            raise HTTPException(status_code=400, detail="Elimina primero las reglas por proveedor para convertir la lista en una lista de compra")
    if "source_id" in update_data:
        await _validate_source(db, update_data["source_id"], current_user.company_id)
    effective_base_source = update_data.get("base_source", pricelist.base_source)
    effective_source_id = update_data.get("source_id", pricelist.source_id)
    if effective_base_source.startswith("EXTERNAL") and not effective_source_id:
        raise HTTPException(status_code=400, detail="Selecciona el origen de datos para usar un precio externo")
    for field, value in update_data.items():
        setattr(pricelist, field, value)
    
    await db.commit()
    refreshed = await db.execute(
        select(PriceList).where(
            PriceList.id == pricelist.id,
            PriceList.company_id == current_user.company_id,
        ).options(selectinload(PriceList.items), selectinload(PriceList.source_rules))
    )
    pricelist = refreshed.scalars().first()
    await log_activity(db, "UPDATE", "PriceList", pricelist.id, current_user.id, current_user.company_id)
    return pricelist

@router.post("/{id}/items", response_model=schemas.PriceListItem)
async def add_pricelist_item(
    *,
    db: AsyncSession = Depends(get_db),
    id: str,
    item_in: schemas.PriceListItemCreate,
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> Any:
    """
    Add item to price list.
    """
    pricelist = await _get_pricelist(db, id, current_user)
    await _validate_item_product(db, item_in, current_user.company_id)
    item_query = select(PriceListItem).where(
        PriceListItem.pricelist_id == pricelist.id,
        PriceListItem.product_id == item_in.product_id,
        PriceListItem.variant_id == item_in.variant_id if item_in.variant_id else PriceListItem.variant_id.is_(None),
    )
    item = (await db.execute(item_query)).scalars().first()
    if item is None:
        item = PriceListItem(pricelist_id=pricelist.id, **item_in.model_dump())
        db.add(item)
    else:
        item.min_quantity = item_in.min_quantity
        item.price = item_in.price
        item.is_active = True
    await db.commit()
    await db.refresh(item)
    return item


@router.put("/{id}/items/{item_id}", response_model=schemas.PriceListItem)
async def update_pricelist_item(
    *,
    db: AsyncSession = Depends(get_db),
    id: str,
    item_id: str,
    item_in: schemas.PriceListItemUpdate,
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> Any:
    await _get_pricelist(db, id, current_user)
    item = await db.scalar(
        select(PriceListItem).where(
            PriceListItem.id == item_id,
            PriceListItem.pricelist_id == id,
        )
    )
    if not item:
        raise HTTPException(status_code=404, detail="Ítem de la lista no encontrado")
    for field, value in item_in.model_dump(exclude_unset=True).items():
        setattr(item, field, value)
    await db.commit()
    await db.refresh(item)
    return item


@router.delete("/{id}/items/{item_id}", status_code=204)
async def delete_pricelist_item(
    *,
    db: AsyncSession = Depends(get_db),
    id: str,
    item_id: str,
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> None:
    await _get_pricelist(db, id, current_user)
    item = await db.scalar(
        select(PriceListItem).where(
            PriceListItem.id == item_id,
            PriceListItem.pricelist_id == id,
        )
    )
    if not item:
        raise HTTPException(status_code=404, detail="Ítem de la lista no encontrado")
    await db.delete(item)
    await db.commit()


@router.post("/{id}/global-adjustment", response_model=schemas.PriceList)
async def apply_global_adjustment(
    *,
    db: AsyncSession = Depends(get_db),
    id: str,
    adjustment: schemas.PriceListGlobalAdjustment,
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> Any:
    pricelist = await _get_pricelist(db, id, current_user)
    if str(getattr(pricelist.type, "value", pricelist.type)) != "SALE":
        raise HTTPException(status_code=400, detail="El aumento global solo aplica a listas de venta")
    pricelist.pricing_mode = "MARKUP_PERCENT"
    pricelist.adjustment_value = adjustment.percent
    if adjustment.base_source is not None:
        pricelist.base_source = adjustment.base_source
    if pricelist.base_source.startswith("EXTERNAL") and not pricelist.source_id:
        raise HTTPException(status_code=400, detail="La lista necesita un origen de datos para calcular precios externos")
    if adjustment.rounding_step is not None:
        pricelist.rounding_step = adjustment.rounding_step
    if adjustment.min_margin_percent is not None:
        pricelist.min_margin_percent = adjustment.min_margin_percent
    if not adjustment.preserve_overrides:
        await db.execute(delete(PriceListItem).where(PriceListItem.pricelist_id == pricelist.id))
    await db.commit()
    result = await db.execute(
        select(PriceList)
        .where(PriceList.id == pricelist.id)
        .options(selectinload(PriceList.items), selectinload(PriceList.source_rules))
    )
    return result.scalars().first()


async def _build_price_export_rows(db: AsyncSession, pricelist: PriceList, company_id) -> list[dict[str, Any]]:
    product_result = await db.execute(
        select(Product)
        .where(Product.company_id == company_id, Product.is_active.is_(True))
        .options(selectinload(Product.variants))
        .order_by(Product.name.asc())
    )
    products = product_result.scalars().all()
    context = await load_price_list_context(db, pricelist.id, [product.id for product in products])
    item_map = context.items
    rows: list[dict[str, Any]] = []
    for product in products:
        variants = list(product.variants or [])
        if not variants:
            variants = [None]
        for variant in variants:
            key = (product.id, variant.id if variant else None)
            external_price, external_cost = selected_external_values(context, product, variant)
            rows.append({
                "PRODUCT_ID": str(product.id),
                "VARIANT_ID": str(variant.id) if variant else "",
                "SKU": (variant.sku if variant and variant.sku else product.sku) or "",
                "VARIANT_NAME": variant.name if variant else "",
                "PRODUCT_NAME": product.name or "",
                "EXTERNAL_PRICE": external_price,
                "EXTERNAL_COST": external_cost,
                "BASE_PRICE": (
                    float(variant.price) if variant and variant.price is not None
                    else float(product.price or 0) + float(variant.price_extra or 0) if variant
                    else float(product.price or 0)
                ),
                "CALCULATED_PRICE": resolve_price(context, product, variant),
                "PRICE_OVERRIDE": item_map.get(key) if key in item_map else item_map.get((product.id, None), ""),
            })
    return rows


@router.get("/{id}/export.xlsx")
async def export_pricelist_excel(
    id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> StreamingResponse:
    pricelist = await _get_pricelist(db, id, current_user)
    rows = await _build_price_export_rows(db, pricelist, current_user.company_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Precios"
    headers = [
        "PRODUCT_ID", "VARIANT_ID", "SKU", "VARIANT_NAME", "PRODUCT_NAME",
        "EXTERNAL_PRICE", "EXTERNAL_COST", "BASE_PRICE", "CALCULATED_PRICE", "PRICE_OVERRIDE",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(48, max(14, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[column[0].column_letter].width = width
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="lista-precios-{id}.xlsx"'},
    )


async def _read_pricelist_excel(file: UploadFile) -> list[dict[str, Any]]:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Sube un archivo Excel .xlsx")
    content = await file.read()
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="El archivo supera el límite de 15 MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel") from exc
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo Excel está vacío")
    headers = [str(value or "").strip().upper() for value in rows[0]]
    aliases = {
        "PRODUCT_ID": {"PRODUCT_ID", "PRODUCTO_ID", "ID_PRODUCTO"},
        "VARIANT_ID": {"VARIANT_ID", "VARIANTE_ID", "ID_VARIANTE"},
        "PRICE_OVERRIDE": {"PRICE_OVERRIDE", "PRECIO_OVERRIDE", "PRECIO_VENTA", "PRECIO_LISTA", "PRICE"},
    }
    positions: dict[str, int] = {}
    for canonical, values in aliases.items():
        for index, header in enumerate(headers):
            if header in values:
                positions[canonical] = index
                break
    if "PRODUCT_ID" not in positions or "PRICE_OVERRIDE" not in positions:
        raise HTTPException(status_code=400, detail="El Excel debe incluir PRODUCT_ID y PRICE_OVERRIDE")
    parsed: list[dict[str, Any]] = []
    for line_number, values in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in values):
            continue
        product_id = str(values[positions["PRODUCT_ID"]] or "").strip()
        variant_id = str(values[positions["VARIANT_ID"]] or "").strip() if "VARIANT_ID" in positions else ""
        parsed.append({
            "line": line_number,
            "product_id": product_id,
            "variant_id": variant_id or None,
            "price": _parse_excel_number(values[positions["PRICE_OVERRIDE"]]),
            "raw_price": values[positions["PRICE_OVERRIDE"]],
        })
    return parsed


@router.post("/{id}/import.xlsx", response_model=schemas.PriceListImportResult)
async def import_pricelist_excel(
    id: str,
    file: UploadFile = File(...),
    dry_run: bool = True,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> schemas.PriceListImportResult:
    pricelist = await _get_pricelist(db, id, current_user)
    rows = await _read_pricelist_excel(file)
    errors: list[str] = []
    valid: list[dict[str, Any]] = []
    parsed_rows: list[dict[str, Any]] = []
    product_ids: set[uuid.UUID] = set()
    variant_ids: set[uuid.UUID] = set()
    for row in rows:
        try:
            product_id = uuid.UUID(row["product_id"])
            variant_id = uuid.UUID(row["variant_id"]) if row["variant_id"] else None
        except (ValueError, TypeError, AttributeError):
            errors.append(f"Fila {row['line']}: identificador de producto o variante inválido")
            continue
        if row["raw_price"] not in (None, "") and row["price"] is None:
            errors.append(f"Fila {row['line']}: precio inválido")
            continue
        parsed = {**row, "product_id_uuid": product_id, "variant_id_uuid": variant_id}
        parsed_rows.append(parsed)
        product_ids.add(product_id)
        if variant_id:
            variant_ids.add(variant_id)

    product_result = await db.execute(
        select(Product.id).where(
            Product.company_id == current_user.company_id,
            Product.id.in_(product_ids) if product_ids else False,
        )
    )
    known_product_ids = set(product_result.scalars().all())
    variant_result = await db.execute(
        select(ProductVariant.id, ProductVariant.product_id).where(
            ProductVariant.company_id == current_user.company_id,
            ProductVariant.id.in_(variant_ids) if variant_ids else False,
        )
    )
    known_variants = {variant_id: product_id for variant_id, product_id in variant_result.all()}
    seen_keys: set[tuple[uuid.UUID, uuid.UUID | None]] = set()
    for row in parsed_rows:
        product_id = row["product_id_uuid"]
        variant_id = row["variant_id_uuid"]
        if product_id not in known_product_ids:
            errors.append(f"Fila {row['line']}: el producto no existe en esta empresa")
            continue
        if variant_id and (variant_id not in known_variants or known_variants[variant_id] != product_id):
            errors.append(f"Fila {row['line']}: la variante no pertenece al producto")
            continue
        key = (product_id, variant_id)
        if key in seen_keys:
            errors.append(f"Fila {row['line']}: el producto y variante están repetidos en el archivo")
            continue
        seen_keys.add(key)
        valid.append(row)

    existing_result = await db.execute(
        select(PriceListItem).where(
            PriceListItem.pricelist_id == pricelist.id,
            PriceListItem.product_id.in_(product_ids) if product_ids else False,
        )
    )
    existing_items = {
        (item.product_id, item.variant_id): item
        for item in existing_result.scalars().all()
    }
    if not dry_run:
        created = updated = cleared = 0
        for row in valid:
            key = (row["product_id_uuid"], row["variant_id_uuid"])
            item = existing_items.get(key)
            if row["price"] is None:
                if item:
                    await db.delete(item)
                    cleared += 1
                continue
            if item:
                item.price = row["price"]
                item.is_active = True
                updated += 1
            else:
                db.add(PriceListItem(
                    pricelist_id=pricelist.id,
                    product_id=row["product_id_uuid"],
                    variant_id=row["variant_id_uuid"],
                    price=row["price"],
                ))
                created += 1
        await db.commit()
    else:
        created = updated = cleared = 0
    return schemas.PriceListImportResult(
        dry_run=dry_run,
        rows_received=len(rows),
        rows_applied=0 if dry_run else len(valid),
        rows_created=created,
        rows_updated=updated,
        rows_cleared=cleared,
        rows_failed=len(errors),
        errors=errors[:100],
    )

@router.delete("/{id}", response_model=schemas.PriceList)
async def delete_pricelist(
    *,
    db: AsyncSession = Depends(get_db),
    id: str,
    current_user: User = Depends(PermissionChecker("manage_inventory")),
) -> Any:
    """
    Delete a price list and its items.
    """
    query = select(PriceList).where(
        PriceList.id == id,
        PriceList.company_id == current_user.company_id
    ).options(selectinload(PriceList.items))
    result = await db.execute(query)
    pricelist = result.scalars().first()
    if not pricelist:
        raise HTTPException(status_code=404, detail="Price list not found")
    
    # Delete items first
    for item in pricelist.items:
        await db.delete(item)
    
    await db.delete(pricelist)
    await db.commit()
    await log_activity(db, "DELETE", "PriceList", id, current_user.id, current_user.company_id)
    return pricelist

